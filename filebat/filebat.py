# -*- coding: utf-8 -*-
# 对文件批量移动、重命名、删除等功能的工具。

# pip install openpyxl
# pip install Gooey
# 解决使用pyinstaller打包后，Gooey中print编码错误：
    # 修改site-packages\gooey\gui\processor.py
    # self.encoding = 'gbk' # encoding

from openpyxl import load_workbook
from gooey import Gooey,GooeyParser

import os
import shutil

# 公用配置字典
CONF = {}

def freemove(fp,new_fp,iscopy=False,is_replace_folder=True):
    '''
    自由移动/复制单一文件(文件夹)至目标文件夹(或重命名文件)
    所经文件夹不存在则递归创建
    不处理无后缀文件
    @param fp: 源文件路径及文件名/文件夹
    @param new_fp: 目标文件名/文件夹
    @iscopy: True为复制，False为移动
    @is_replace_folder: 文件夹冲突时，清空原文件夹还是保留原内容
    '''
    if os.path.isdir(fp):   # 左->文件夹(右->只能是文件夹)
        if os.path.exists(new_fp):
            if is_replace_folder:
                shutil.rmtree(new_fp)
                print(f'old folder <{fp}> deleted!')
            else:   # 暂不做递归覆盖
                print(f'folder <{fp}> exists, do nothing!')
                return
        if not new_fp:  # 只填左侧，右侧留空则为左侧删除标记
            shutil.rmtree(fp)
            print(f'deleted(folder): {fp}')
            return
        if iscopy:
            shutil.copytree(fp,new_fp)
            print(f'folder copied: {fp} --> {new_fp}')
        else:
            shutil.move(fp,new_fp)
            print(f'folder moved: {fp} --> {new_fp}')
    else:   # 左->文件
        if not os.path.exists(fp):
            if not fp and new_fp and not os.path.exists(new_fp):    # 左侧留空，只填右侧为批量创建文件夹
                os.makedirs(new_fp)
                print(f'created: {new_fp}')
            else:
                print(f'ERROR: {fp} not exists!')
            return
        if not new_fp:    # 只填左侧，右侧留空则为左侧删除标记
            os.remove(fp)
            print(f'deleted: {fp}')
            return
        ext = os.path.splitext(fp)[1]
        new_ext = os.path.splitext(new_fp)[1]
        if not ext:   # 无后缀文件不处理
            print(f'ERROR: {fp} has no extension, do nothing!')
            return
        elif not new_ext:   # 右->文件夹
            new_path = new_fp
        else:   # 右->文件(文件名及后缀可同可不同)
            new_path = os.path.split(new_fp)[0]
        if new_path and not os.path.exists(new_path):    # 递归补建文件夹
            os.makedirs(new_path)
        if iscopy:
            shutil.copy(fp,new_fp)
            print(f'copied: {fp} --> {new_fp}')
        else:
            shutil.move(fp,new_fp)
            print(f'moved: {fp} --> {new_fp}')

def load_config(fp):
    '''
    从Excel文件中加载配置至公用配置字典CONF
    '''
    fp = os.path.normcase(fp)
    if os.path.exists(fp):
        wb = load_workbook(filename = fp)
        CONF['rLEFT_FILES'] = fetch_name(wb,'rLEFT_FILES')
        CONF['rRIGHT_FILES'] = fetch_name(wb,'rRIGHT_FILES')  
        CONF['IS_REPLACE_FOLDER'] = fetch_name(wb,'IS_REPLACE_FOLDER')
        CONF['IS_COPY'] = fetch_name(wb,'IS_COPY')
        assert len(CONF['rLEFT_FILES']) == len(CONF['rRIGHT_FILES'])    # 左右文件清单长度需一致
        
def fetch_name(wb,defined_name,is_return_cell=False):
    '''
    从当前的Excel文件中，加载名称对应的数据
    @param wb: 使用openpyxl加载Excel的workbook
    @param defined_name: Excel中定义的名称
    @param is_return_cell: 是否返回cell对象
    @return: cell对象或列表；泛型数值或列表
    '''
    dn = wb.defined_names[defined_name]
    cells = []
    for k,v in dn.destinations:
        ws = wb[k]
        cells.append(ws[v])
    cells = cells[0]    # 去掉无用的列表层
    # 返回cell对象
    if is_return_cell:
        return cells
    # 返回值
    if type(cells) is tuple: # 区域
        return [cell.value for cell in cells][5:]   # Excel配置前五行为说明
    else:   # 单独单元格
        return cells.value

@Gooey(program_name = '批量文件操作工具', language = 'chinese')
def main():
    parser = GooeyParser(description = '对文件批量移动、重命名、删除等功能的工具。\n提示：操作无法撤销，请注意备份！')
    parser.add_argument('conf_file', help = '配置文件，默认无需修改', widget = 'FileChooser', default = 'filebat.xlsx') 
    args = parser.parse_args()
    # 加载配置
    load_config(args.conf_file)
    total_count = len(CONF['rLEFT_FILES'])
    print(f'【filebat】配置已加载，待处理{total_count}项内容')
    # 批量处理
    for num in range(total_count):
        fn = CONF['rLEFT_FILES'][num] or ''
        new_fn = CONF['rRIGHT_FILES'][num] or ''
        if fn == '' and new_fn == '':continue
        print(f'【{num+1}】', end = '')
        freemove(fp = fn, 
                 new_fp = new_fn,
                 iscopy = CONF['IS_COPY'],
                 is_replace_folder = CONF['IS_REPLACE_FOLDER'])
    print('【filebat】全部处理完毕！')
    
if __name__ == '__main__':
    main()
