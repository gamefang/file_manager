# -*- coding: utf-8 -*-

import os
import time
from openpyxl import load_workbook

from DataManager import DataManager as DataManager

class XlManager():
    '''
    实现Excel读写的功能
    '''
    # 当前加载的Excel文件
    CUR_WB = None

    @classmethod
    def load_cur_file(cls,fp):
        '''
        加载当前使用的Excel文件，供随时访问
        @param fp: Excel文件路径
        '''
        fp = os.path.normcase(fp)
        if os.path.exists(fp):
            cls.CUR_WB = load_workbook(filename = fp)

    @classmethod
    def fetch_name(cls,defined_name,is_return_cell=False):
        '''
        从当前的Excel文件中，加载名称对应的数据
        @param defined_name: Excel中定义的名称
        @param is_return_cell: 是否返回cell对象
        @return: cell对象或列表；泛型数值或列表
        '''
        if not cls.CUR_WB:return
        dn = cls.CUR_WB.defined_names[defined_name]
        cells = []
        for k,v in dn.destinations:
            ws = cls.CUR_WB[k]
            cells.append(ws[v])
        cells = cells[0]    # 去掉无用的列表层
        # 返回cell对象
        if is_return_cell:
            return cells
        # 返回值
        if type(cells) is tuple: # 区域
            return [ cell.value
                for cell in cells
                if cell.value
                ][1:]
        else:   # 单独单元格
            return cells.value

    @classmethod
    def load_excel_data(cls,CFG):
        '''
        加载Excel文件中的数据
        @param CFG: 配置信息，来自ConfManager
        @return: 文件数据字典
        '''
        workbook = load_workbook(filename = CFG.BASE['EXCEL_FILE_PATH'], read_only=True)
        sheet_name = CFG.BASE['LIST_SHEET_NAME']
        if sheet_name not in workbook.sheetnames:
            raise Exception('Excel file wrong!')
        ws = workbook[sheet_name]
        # 先快速从表格读取数据为列表
        data = []
        is_start = False
        for row in ws.rows:
            cur_list = []
            for cell in row:
                if not is_start and cell.value == 'key':
                    is_start = True
                cur_list.append(cell.value)
            if is_start:
                data.append(cur_list)
        workbook.close()
        # 列表转为字典
        excel_data = {}
        headers = data[0]
        for row in data[2:]:
            cur_dic = {}
            for num,value in enumerate(row):
                # 跳过key以及没有值的单元格
                if num == 0 or value is None:continue
                cur_dic[headers[num]] = value
            # 确定key
            list_keys = excel_data.keys()
            if row[0]:
                cur_key = row[0]
            else:
                cur_key = None
            key = DataManager.get_key(cur_dic,CFG,list_keys,cur_key)
            # 数据记录
            excel_data[key] = cur_dic
        # 强制备份Excel数据
        DataManager.backup_data(excel_data)
        return excel_data

    @classmethod
    def write_data_to_excel(cls,data,CFG):
        '''
        写入最终数据至Excel文件中
        @param data: 待写入data（来自FileManager的处理结果）
        @param CFG: 配置信息，来自ConfManager
        '''
        workbook = cls.CUR_WB
        sheet_name = CFG.BASE['LIST_SHEET_NAME']
        if sheet_name not in workbook.sheetnames:
            raise Exception('Excel file wrong!')
        # sheet备份（自动加顺序号）
        if CFG.EXCEL['AUTO_BACKUP']:
            sheet_copy = workbook.copy_worksheet(workbook[sheet_name])
        # 使用原sheet
        ws = workbook[sheet_name]
        # 表头提取
        head_start_cell = cls.get_cell_by_value(ws,'key')
        head_row = ws[head_start_cell.row]
        # 根据表头确定输出字段顺序
        output_params = [cell.value for cell in head_row]
        # 清空无用表格行（保留表头下方文字表头行）
        ws.delete_rows(head_start_cell.row + 2, ws.max_row)
        # 按顺序输出
        p_row = head_start_cell.row + 2 # row指针
        for k,v in data.items():
            p_col = head_start_cell.column # col指针
            for item in output_params:
                # 输出key
                if item == 'key':
                    cur_v = k
                # 自动套用公式
                elif item == 'hyperlink':
                    ref_cell_addr = ws.cell(p_row,output_params.index('path')+1).coordinate
                    cur_v = f'=HYPERLINK(BASE_FOLDER&"/"&{ref_cell_addr},"打开")'
                elif item == 'filetype':
                    ref_cell_addr = ws.cell(p_row,output_params.index('ext')+1).coordinate
                    cur_v = f'=IFERROR(VLOOKUP({ref_cell_addr},rEXT_TO_TYPE,2,),"")'
                # 时间戳处理
                elif item in ('c_time','m_time','a_time'):
                    cur_timestamp = v.get(item.replace('_',''))
                    if cur_timestamp:
                        cur_v = cls.timestamp_to_str(cur_timestamp)
                    else:
                        cur_v = ''
                # 正常输出各字段
                else:
                    cur_v = v.get(item,'')  # 留空不存在数据
                cur_cell = ws.cell(
                    column = p_col,
                    row = p_row,
                    value = cur_v,
                )
                p_col += 1
            p_row += 1
        # 存储
        workbook.save(CFG.BASE['EXCEL_FILE_PATH'])

    @staticmethod
    def get_cell_by_value(sheet,value):
        '''
        按值获取单元格对象（先横后纵，取第一个值）
        @param sheet: 值所在的Excel工作表
        @param value: 所需的值
        @return: cell对象
        '''
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == value:
                    return cell

    @staticmethod
    def is_excel_opened(fp):
        '''
        判断Excel文件是否已打开（通过是否生成了~$文件判定）
        @param fp: Excel文件路径
        @return: bool
        '''
        fp = os.path.normcase(fp)
        dir_name,file_name = os.path.split(fp)
        hidden_fp = os.path.join(dir_name,'~$' + file_name)
        return os.path.exists(hidden_fp)

    @staticmethod
    def timestamp_to_str(timestamp):
        '''
        时间戳转字符串时间
        @param timestamp: 时间戳
        @return: 字符串时间
        '''
        time_array = time.localtime(timestamp)
        return time.strftime('%Y-%m-%d %H:%M:%S',time_array)


if __name__ == '__main__':
    XlManager.load_cur_file('FileManager.xlsx')
    res = XlManager.fetch_name('NO_HIDDEN_FILES_WIN')
    print(res)
    res = XlManager.fetch_name('rEXT_WHITELIST')
    print(res)
